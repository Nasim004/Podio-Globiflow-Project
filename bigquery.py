import os
import json
import requests
import google.auth
import openpyxl
from google.cloud import bigquery
from podio_api import get_access_token

USERNAME = os.environ["string_1"]
PASSWORD = os.environ["string_2"]
sheet = "sheets/Columns.xlsx"

def get_assets(assets) ->tuple: 
    # this is for to compare the corresponding assets in the dict (assets is key here)
    asset_key_value = {'Commercial (General)': ['Commercial (General)', 'Null'], 'Self-Storage': ['Self-Storage', 'Public Storage'], 'Commercial Condominium': ['Commercial Condominium'], 'Mobile Home Park': ['Mobile Home Park', 'RV Park'], 'Multi Family (General)': ['Multi Family (General)'], 'Cooperative': ['Cooperative'], 'Multi-Family 2-4 Unit': ['Quadruplex', 'Duplex', 'Triplex'], 'Nursing Home': ['Nursing Home'], 'Dormitories/Group Quarters': ['Dormitories/Group Quarters'], 'Frat/Sorority House': ['Frat/Sorority House'], 'Warehouse': ['Warehouse'], 'Storage': ['Storage'], 'General Industrial': ['Light Industrial', 'Heavy Industrial', 'General Industrial'], 'Industrial Condominium': ['Industrial Condominium'], 'Chemical': ['Chemical'], 'Shipyard': ['Shipyard'], 'Industrial Park': ['Industrial Park', 'Industrial Plant'], 'Mine/Quarry': ['Mine/Quarry'], 'Lumber Mill': ['Lumber Yard', 'Lumber Mill'], 'Metal Production': ['Metal Production'], 'Aircraft Facility': ['Aircraft Facility'], 'Paper & Allied Industry': ['Paper & Allied Industry'], 'Food Packing/Packing/Canning Plant': ['Food Packing/Packing/Canning Plant'], 'Storage Tanks': ['Storage Tanks'], 'Salvage': ['Salvage'], 'Gas Production': ['Gas Production', 'Petroleum'], 'Mineral Processing': ['Mineral Processing'], 'Transportation': ['Transportation'], 'Waste Disposal': ['Waste Disposal', 'Dump Site'], 'Automotive Wrecking': ['Automotive Wrecking'], 'Textile/Clothes/Carpet': ['Textile/Clothes/Carpet'], 'Facilities': ['Facilities'], 'Bulk Plant': ['Bulk Plant']}
    # getting the value and convert them into tuple
    final_asset_type = set(item for key in assets for item in asset_key_value[key])
    # to avoid last comma if set has only 1 element
    final_asset_type.add("None")
    final_asset_type = tuple(final_asset_type)
    return final_asset_type

def bigquery_data( address, asset_type):
    credentials, project = google.auth.default(
        scopes=[
            "https://www.googleapis.com/auth/drive",
            "https://www.googleapis.com/auth/cloud-platform",
        ]
    )
    client = bigquery.Client(credentials=credentials, project=project)  
    # getting the zipcode from address , eg:" 3 E 64th St, New York, NY 10065, USA "
    zip_code = address.split(",")
    zip_code = zip_code[-2]
    zip_code = zip_code.split(" ")[-1]
    print("Zip Code ======> ",zip_code)
    # this is for copying the column names from the Column.xlsx to new sheet
    data_frame = openpyxl.load_workbook(sheet)
    # sheet 1 selecting
    worksheet = data_frame.active
    header_row_1 = worksheet[1]
    # sheet 2 selecting
    worksheet = data_frame["buyer_type_cash"]
    header_row_2 = worksheet[1]

    upload_file = openpyxl.Workbook()
    agent_sheet = upload_file.active
    agent_sheet.title = "Agents"

    upload_file.create_sheet(title="buyer_type_cash")
    cash_sheet = upload_file["buyer_type_cash"]
    for cell in header_row_2:
        cash_sheet.cell(row=1, column=cell.column, value=cell.value) 
        
    final_assets = get_assets(asset_type)    
    counties = set()
    
    query_job = client.query(
    f"""
    SELECT Primary_Official_County_Name 
    FROM `reia-commercial-dealflyte.county_zips.county_zip_detail` 
    WHERE Zip_Code = {zip_code}  
    """
    )
    results = query_job.result()
    for result in results:
      result = result.values()
      counties.add(result[0])
    # to avoid last comma if set has only 1 element
    counties.add("None")
    counties = tuple(counties)
    print("Counties ======> ",counties)    
    print(f"""
    SELECT *
    FROM `reia-commercial-dealflyte.commercial_buyers.master_buyers_1`
    JOIN `reia-commercial-dealflyte.commercial_buyers.master_buyer_contacts`
    ON `reia-commercial-dealflyte.commercial_buyers.master_buyers_1`.unique_id = `reia-commercial-dealflyte.commercial_buyers.master_buyer_contacts`.unique_id 
    WHERE asset_detail_category IN {final_assets} AND county IN {counties} AND buyer_type = 'Cash/Conventional' 
    """)

    print("Cash Query Started")
    
    query_job = client.query(
    f"""
    SELECT *
    FROM `reia-commercial-dealflyte.commercial_buyers.master_buyers_1`
    JOIN `reia-commercial-dealflyte.commercial_buyers.master_buyer_contacts`
    ON `reia-commercial-dealflyte.commercial_buyers.master_buyers_1`.unique_id = `reia-commercial-dealflyte.commercial_buyers.master_buyer_contacts`.unique_id 
    WHERE asset_detail_category IN {final_assets} AND county IN {counties} AND  buyer_type = 'Cash'
    """
    )
    
    results = query_job.result()
    for row in results:
        row_values = list(row.values())
        cash_sheet.append(row_values)
    
    cash_sheet.delete_cols(16,1)
    upload_file.create_sheet(title="buyer_type_conventional")  
    conventional_sheet = upload_file["buyer_type_conventional"]
    for cell in header_row_2:
        conventional_sheet.cell(row=1, column=cell.column, value=cell.value)
    print("Conventional Query Started")
    query_job = client.query(
    f"""
    SELECT *
    FROM `reia-commercial-dealflyte.commercial_buyers.master_buyers_1`
    JOIN `reia-commercial-dealflyte.commercial_buyers.master_buyer_contacts`
    ON `reia-commercial-dealflyte.commercial_buyers.master_buyers_1`.unique_id = `reia-commercial-dealflyte.commercial_buyers.master_buyer_contacts`.unique_id 
    WHERE asset_detail_category IN {final_assets} AND county IN {counties} AND buyer_type = 'Conventional'
    """
    )
    results = query_job.result()
    for row in results:
        row_values = list(row.values())
        conventional_sheet.append(row_values)
    
    conventional_sheet.delete_cols(16,1)
    # copying the column name
    for cell in header_row_1:
        agent_sheet.cell(row=1, column=cell.column, value=cell.value)
    print("Agent Query Started")
    
    # query to retrieve the data from master_agents sheet
    print(f"SELECT * FROM commercial_agents.master_agents WHERE county IN {counties}")
    query_job = client.query(
        f"SELECT * FROM commercial_agents.master_agents WHERE county IN {counties} "
    )
    results = query_job.result()
    for row in results:
        row_values = list(row.values())
        row_values.insert(12,f"{zip_code}")
        agent_sheet.append(row_values)
    upload_file.save("Upload File.xlsx")
    return None

def file_upload(item_id):
    BASE_URL = 'https://api.podio.com/'
    # uploading the file to podio
    token = get_access_token(USERNAME,PASSWORD)
    print('file upload started') 
    url = BASE_URL + "file/"
    payload = {'filename': 'interestedinvestor.xlsx'}
    files=[('source',(f'Upload File.xlsx',open(f'Upload File.xlsx','rb'),'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))]
    headers = {
    'Authorization': f'Bearer {token}'
    }
    response = requests.post( url, headers=headers, data=payload, files=files)
    file_id = response.json()['file_id']
    if file_id:
        print("==== File Uploaded ====")
    # adding the file to podio item
    url = BASE_URL + f"file/{file_id}/attach"
    id = item_id
    payload = json.dumps({
            "ref_type": "item",
            "ref_id": id
            })
    response = requests.post(url,headers=headers,data=payload)
    print(response.text)
    if response.status_code == 204:
        print("===== File Added =====")


