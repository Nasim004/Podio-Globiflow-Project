import os
import json
import time
import shutil
import requests
import clicksend_client
from dotenv import load_dotenv
from openpyxl import load_workbook
from podio_api import get_access_token
from selenium_function import file_download
from clicksend_client.rest import ApiException
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service

load_dotenv()
username_podio = os.environ["string_1"]
password_podio = os.environ["string_2"]
username_click_send = os.environ["string_3"]
password_click_send = os.environ["string_4"]


def get_file_link(token, item_id):
    url = f"https://api.podio.com/item/{item_id}"
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        data = response.json()
        for id in data["files"]:
            if id["name"] == "interestedinvestor.xlsx":
                file_id = id["file_id"]
                url = f"https://api.podio.com/file/{file_id}"
                response = requests.get(url, headers=headers)
                if response.status_code == 200:
                    data = response.json()
                    return data["link"]


def list_id(username, api_key):
    configuration = clicksend_client.Configuration()
    configuration.username = username
    configuration.password = api_key
    # create an instance of the API class
    api_instance = clicksend_client.ContactListApi(
        clicksend_client.ApiClient(configuration)
    )
    list = clicksend_client.ContactList(list_name="Podio List")  # List | List model
    try:
        # Create new contact list
        api_response = api_instance.lists_post(list)
        api_response = api_response.replace("'", '"')
        api_response_dict = json.loads(api_response)
        return api_response_dict["data"]["list_id"]
    except ApiException as e:
        print("Exception when calling ContactListApi->lists_post: %s\n" % e)
# This is get API key from CLick Send
# Login by username and password then get the api 
def get_api_key(username, password):
    print("API Key")
    options = webdriver.ChromeOptions()
    # options.add_argument("--headless")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("enable-automation")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-dev-shm-usage")
    service = Service(executable_path="chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=options)
    driver.get("https://dashboard.clicksend.com/login")
    elem = driver.find_element(By.NAME, "username")
    elem.send_keys(username)
    time.sleep(1)
    elem = driver.find_element(By.NAME, "password")
    elem.send_keys(password)
    time.sleep(1)
    driver.find_element(By.NAME, "login").click()
    time.sleep(15)
    driver.find_element(By.NAME, "api-credentials").click()
    time.sleep(2)
    api_key = driver.find_element(By.XPATH,"/html/body/div[1]/div/div[1]/nav/div/div[2]/div/div/div[2]/div/div[2]/div[2]").text
    return api_key

def buyer_sheet(sheet_name, api_instance, id):
    dataframe = load_workbook("interestedinvestor.xlsx")
    sheet = dataframe[sheet_name]
    i = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[46] is None:
            continue
        if row[54] is None:
            email = "noemail@gmail.com"
        else:
            email = row[54]
        contact = clicksend_client.Contact( phone_number=row[46], first_name=row[2], custom_1="-", email=email)
        try:
            api_response = api_instance.lists_contacts_by_list_id_post(contact, id)
            i += 1
        except ApiException as e:
            print("Exception when calling ContactApi->lists_contacts_by_list_id_post: %s\n",e)
            continue
    # for to check how many api are done
    print("Data Passed ===>",i)

def contact_data(username, password, option):
    print("Inside contact data")
    # to get the api key
    api_key = get_api_key(username, password)
    print("API Key =====>", api_key)
    # moving the downloaded file to project folder
    # for local server
    shutil.move(
        "C:\\Users\\Nasim\\Downloads\\interestedinvestor.xlsx",
        "D:\\ODZ Services\\Projects\\podio",
    )
    # for server
    # removing the file so the is no erro occur when moving the file
    # shutil.move('C:\\Users\\administratorpodio\\Downloads\\interestedinvestor.xlsx', 'C:\\Users\\administratorpodio\\Documents\\Podio')
    sheet = "interestedinvestor.xlsx"
    dataframe = load_workbook(sheet)
    worksheet = dataframe.active
    # calling click send contact create api
    # Configure HTTP basic authorization: BasicAuth
    configuration = clicksend_client.Configuration()
    configuration.username = username
    configuration.password = api_key
    # create an instance of the API class
    api_instance = clicksend_client.ContactApi(clicksend_client.ApiClient(configuration))
    if option == "1":
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # Contact model
            if row[25] is None:
                custom_1_value = "None"
            else:
                custom_1_value = row[25]
            if row[23] is None:
                continue
            phone = row[3].split(":")[-1]
            if email is None:
                email = "None"
            else:
                email = row[2].split(":")[-1]
            contact = clicksend_client.Contact(
                phone_number=phone,
                first_name=row[0],
                custom_1=custom_1_value,
                custom_2=row[27],
                email=email,
            )
            # list id for contacts
            id = list_id(username, api_key)
            try:
                # Create new contact
                api_response = api_instance.lists_contacts_by_list_id_post(contact, id)
            except ApiException as e:
                print("Exception when calling ContactApi->lists_contacts_by_list_id_post: %s\n",e)
                continue
    elif option == "2":
        print("inside option 2")
        id = list_id(username, api_key)
        i = 0
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[22] is None:
                phone = 2121231234
            else:
                phone = row[22]
            contact = clicksend_client.Contact(
                phone_number=row[22],
                first_name=row[4],
                custom_1="-",
            )
            try:
                api_response = api_instance.lists_contacts_by_list_id_post(contact, id)
                i += 1
            except:
                # print("Exception when calling ContactApi->lists_contacts_by_list_id_post: %s\n",e)
                continue
        print(i)
        print("sheet 2 started =========>")
        buyer_sheet("buyer_type_cash", api_instance, id)
        print("sheet 3 started =========>")
        buyer_sheet("buyer_type_conventional", api_instance, id)

def click_send_function(item_id, username, password, option):
    # getting token for to get the file link from podio
    token = get_access_token(username_podio, password_podio)
    # getting file link of the item
    file_link = get_file_link(token, item_id)
    print("File Link =======>", file_link)
    # passing the file link to download
    # auth req for downloading the file that's why using selenium method
    # cookies passing through requests is not possible
    file_download(file_link)
    # calling contact creation function
    contact_data(username, password, option)
