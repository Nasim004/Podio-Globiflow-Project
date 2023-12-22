import openpyxl

sheet = "sheets/Zips for Podio V3.xlsx"
def sheet_data_extraction(county):
    dataframe = openpyxl.load_workbook(sheet)
    worksheet = dataframe.active
    cities = set()
    section = set()
    zip_code = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if county in row[4]:
            cities.add(row[1])
            section.add(row[2])
            zip_code.add(row[3])
    return (list(cities), list(section), list(zip_code))

# to get the counties in the xlsx sheet
def get_county():
    data_frame = openpyxl.load_workbook(sheet)
    worksheet = data_frame.active
    counties = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        counties.add(row[4])
    return list(counties)

